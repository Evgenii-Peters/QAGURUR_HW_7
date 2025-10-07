import os
import zipfile
import csv
from pypdf import PdfReader
from openpyxl import load_workbook




def create_zip_files():

    if not os.path.exists("resources"):
        os.mkdir("resources")
    files = [os.path.join('files/', "Example_xlsx.xlsx"),os.path.join('files/', "Example_pdf.pdf"),os.path.join('files/', "Example_csv.csv")]
    with zipfile.ZipFile(os.path.join("resources", "zip_Example.zip"), 'w') as zf:
        for file in files:
            zf.write(file, os.path.basename(file))

create_zip_files()


def test_data_in_csv():
    with zipfile.ZipFile(os.path.join("resources", "zip_Example.zip")) as zf:
        with zf.open("Example_csv.csv") as csv_file:
            content = csv_file.read().decode('utf-8')
            csvreader = list(csv.reader(content.splitlines()))
            last_row = csvreader[-1]
            assert last_row == ['173', '170', '26']

def test_data_in_pdf():
    with zipfile.ZipFile(os.path.join("resources", "zip_Example.zip")) as zf:
        with zf.open("Example_pdf.pdf") as File_pdf:
            reader = PdfReader(File_pdf)
            number_of_pages = len(reader.pages)
            first_page = reader.pages[0].extract_text()

        assert  "Minecraft and MCPACK Files" in first_page
        assert number_of_pages == 9

def test_data_in_xls():
    with zipfile.ZipFile(os.path.join("resources", "zip_Example.zip")) as zf:
        with zf.open("Example_xlsx.xlsx") as File_xlsx:
            workbook = load_workbook(File_xlsx)
            active_sheet = workbook.active
            max_row_sheet = active_sheet.max_row
            first_cell = active_sheet.cell(row=1, column=1).value

        assert max_row_sheet == 28
        assert first_cell == "Ели мясо мужики"






